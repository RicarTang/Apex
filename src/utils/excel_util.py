import os, aiofiles, time, asyncio
from collections import namedtuple
from typing import List, NamedTuple
from pathlib import Path
from fastapi import UploadFile
from openpyxl import load_workbook, Workbook


async def read_all_testcase(filepath: Path) -> List[NamedTuple]:
    """读取excel所有数据

    Args:
        filepath (_type_): excel文件路径

    Returns:
        List[NamedTuple]: _description_
    """
    # 异步循环运行同步函数load_workbook
    loop = asyncio.get_running_loop()
    try:
        wb: Workbook = await loop.run_in_executor(
            None,
            lambda: load_workbook(filepath, data_only=True, read_only=True),
        )
    except Exception as e:
        raise ValueError(f"Failed to load Excel file: {filepath}. Error: {e}")
    # 获取所有sheet
    sheets = wb.sheetnames
    testcases = []  # 存放所有读取的testcase
    # 使用namedtuple
    Testcase: NamedTuple = namedtuple(
        "Testcase",
        [
            "case_no",
            "case_title",
            "case_description",
            "case_module",
            "case_sub_module",
            "case_is_execute",
            "api_path",
            "api_method",
            "request_headers",
            "request_param_type",
            "request_param",
            "expect_code",
            "expect_result",
            "expect_data",
            "request_to_redis",
            "response_to_redis",
            "case_editor",
            "remark",
        ],
    )
    # 把循环读取的耗时任务放进线程池运行避免阻塞主线程
    tasks = [
        loop.run_in_executor(
            None,
            lambda sheet=wb[sheetname]: [
                row for row in sheet.iter_rows(values_only=True) if any(row)
            ],
        )
        for sheetname in wb.sheetnames
    ]
    # 同步读取所有sheet
    all_sheets_data = await asyncio.gather(*tasks)
    testcases = []

    for rows in all_sheets_data:
        testcases.extend(
            [Testcase(*row) for row in rows[1:] if len(row) == len(Testcase._fields)]
        )

    return testcases


async def save_file(file: UploadFile, save_path: Path) -> None:
    """保存上传的文件

    Args:
        file (bytes):  二进制文件
        save_path (Path): 保存的路径
    """
    # 获取文件对象
    uploaded_file = file.file
    # 判断目录
    if not save_path.parent.exists():
        os.makedirs(save_path.parent)
    # 将文件内容写入目标文件
    async with aiofiles.open(save_path, "wb") as f:
        # 避免一次读取内存过大
        while chunk := uploaded_file.read(8192):
            await f.write(chunk)


if __name__ == "__main__":
    path = Path(__file__)
    ex = path.parent.parent.parent / "static" / "testcase" / "测试用例模板.xlsx"

    async def read():
        data = await read_all_testcase(ex)
        print(data)

    async def write():
        file = open(ex, "rb")
        await save_file(file, ex.parent / "upload" / f"{time.time()}.xlsx")

    asyncio.run(read(), debug=True)
